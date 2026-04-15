"""
excel_logger.py — Observabilité Excel FXBOT EUR/USD (M10).
Écriture non-bloquante via asyncio.to_thread.

# MIGRATION-EURUSD : Migration complète XAU/USD → EUR/USD (2026-04-15)
"""

import os
import datetime
import statistics
import math
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXCEL_PATH = "fxbot_eurusd_performance.xlsx"  # MIGRATION-EURUSD : goldbot_performance.xlsx → fxbot_eurusd_performance.xlsx

# ── Colonnes ────────────────────────────────────────────────────────────────

_TRADES_COLS = [
    "date", "heure", "direction", "lot", "entry", "sl", "tp",
    "exit", "pnl_usd", "pnl_pct", "rr_réel", "conf_ia", "source", "durée_min"
]

_SESSION_COLS = [
    "date", "equity", "drawdown_pct", "win_rate",
    "profit_factor", "trades_count", "sharpe_approx"
]

_REFUSED_COLS = [
    "date", "heure", "raison_rejet", "conf", "rr", "source"
]

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)


# ── Init workbook ────────────────────────────────────────────────────────────

def _init_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, cols in [
        ("trades", _TRADES_COLS),
        ("session", _SESSION_COLS),
        ("signaux_refusés", _REFUSED_COLS),
    ]:
        ws = wb.create_sheet(sheet_name)
        for i, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=i, value=col)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(i)].width = max(14, len(col) + 4)

    return wb


def _load_or_create() -> openpyxl.Workbook:
    if os.path.exists(EXCEL_PATH):
        try:
            return openpyxl.load_workbook(EXCEL_PATH)
        except Exception:
            pass
    return _init_workbook()


def _save(wb: openpyxl.Workbook) -> None:
    wb.save(EXCEL_PATH)


# ── Calculs de session ───────────────────────────────────────────────────────

def _compute_session_metrics(wb: openpyxl.Workbook) -> dict:
    ws = wb["trades"]
    pnl_list = []
    wins = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        pnl = row[8]  # pnl_usd
        if pnl is None:
            continue
        pnl_list.append(float(pnl))
        if float(pnl) > 0:
            wins += 1

    total = len(pnl_list)
    win_rate = round(wins / total * 100, 1) if total else 0.0
    positives = [p for p in pnl_list if p > 0]
    negatives = [abs(p) for p in pnl_list if p < 0]
    profit_factor = round(sum(positives) / sum(negatives), 2) if negatives else 0.0

    # Sharpe approx (daily returns grouping)
    sharpe = 0.0
    if len(pnl_list) >= 2:
        try:
            mean = statistics.mean(pnl_list)
            std = statistics.stdev(pnl_list)
            sharpe = round((mean / std) * math.sqrt(252), 2) if std else 0.0
        except Exception:
            pass

    return {
        "win_rate": win_rate,
        "profit_factor": profit_factor,
        "trades_count": total,
        "sharpe_approx": sharpe,
    }


# ── Fonctions d'écriture (synchrones — appelées via asyncio.to_thread) ───────

def write_trade(trade: dict, entry_price: float, exit_price: float,
                pnl_usd: float, source: str, duration_min: int) -> None:
    """Ajoute une ligne dans la feuille 'trades' après clôture."""
    wb = _load_or_create()
    ws = wb["trades"]
    now = datetime.datetime.now()
    sl = trade.get("sl", 0)
    tp = trade.get("tp", 0)

    rr_reel = 0.0
    if sl and tp and entry_price:
        risk = abs(entry_price - sl)
        reward = abs(exit_price - entry_price)
        rr_reel = round(reward / risk, 2) if risk else 0.0

    pnl_pct = round(pnl_usd / trade.get("lot", 1) / entry_price * 100, 3) if entry_price else 0.0

    ws.append([
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M"),
        trade.get("dir", ""),
        trade.get("lot", 0),
        round(entry_price, 2),
        sl,
        tp,
        round(exit_price, 2),
        round(pnl_usd, 2),
        pnl_pct,
        rr_reel,
        trade.get("conf", 0),
        source,
        duration_min,
    ])
    _save(wb)


def write_refused_signal(signal: dict, reason: str, source: str = "deepseek") -> None:
    """Ajoute une ligne dans la feuille 'signaux_refusés'."""
    wb = _load_or_create()
    ws = wb["signaux_refusés"]
    now = datetime.datetime.now()
    ws.append([
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M"),
        reason[:80],
        signal.get("CONF", 0),
        signal.get("RR", 0),
        source,
    ])
    _save(wb)


def update_session_row(equity: float, drawdown_pct: float) -> None:
    """Ajoute/met à jour la ligne de session (toutes les 30 min)."""
    wb = _load_or_create()
    ws = wb["session"]
    metrics = _compute_session_metrics(wb)
    now = datetime.datetime.now()
    ws.append([
        now.strftime("%Y-%m-%d %H:%M"),
        round(equity, 2),
        round(drawdown_pct, 2),
        metrics["win_rate"],
        metrics["profit_factor"],
        metrics["trades_count"],
        metrics["sharpe_approx"],
    ])
    _save(wb)
